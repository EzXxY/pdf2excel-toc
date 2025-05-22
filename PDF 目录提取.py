import sys
import re
import pdfplumber
import os
import openpyxl
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
                             QFileDialog, QMessageBox, QHeaderView, QTextEdit, QMenu, QSizePolicy,
                             QProgressBar, QCheckBox, QInputDialog, QDialog)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QIcon  # 添加QIcon导入
# 添加用于多栏识别的库
import fitz  # PyMuPDF
import io
from PIL import Image
try:
    import pytesseract
    HAS_TESSERACT = True
except ImportError:
    HAS_TESSERACT = False
# 在import部分之后添加
if HAS_TESSERACT:
    try:
        # 设置Tesseract默认路径
        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        # 解除PIL的图像大小限制
        Image.MAX_IMAGE_PIXELS = None
    except:
        pass

def convert_to_chinese_num(n):
    chinese_nums = ['零', '一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
    if isinstance(n, str):
        try:
            n = int(n)
        except ValueError:
            return n
            
    if not isinstance(n, int) or n < 0 or n > 99:
        return str(n)
        
    if n < 11:
        return chinese_nums[n]
    elif n < 20:
        return '十' + (chinese_nums[n % 10] if n % 10 != 0 else '')
    else:
        return chinese_nums[n // 10] + '十' + (chinese_nums[n % 10] if n % 10 != 0 else '')

def convert_roman_to_int(s):
    """将罗马数字转换为整数"""
    roman_dict = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100, 'D': 500, 'M': 1000}
    result = 0
    s = s.upper()
    for i in range(len(s)):
        if i > 0 and roman_dict[s[i]] > roman_dict[s[i-1]]:
            result += roman_dict[s[i]] - 2 * roman_dict[s[i-1]]
        else:
            result += roman_dict[s[i]]
    return result

class OutlineExtractor:
    def __init__(self):
        self.level_configs = []
        self.space_required = []  # 存储每个层级是否需要空格匹配
        self.remove_page_numbers = True  # 新增：控制是否移除页码
        self.colon_truncate = True  # 新增：控制是否在冒号处截断
        self.blocked_keywords = set()  # 新增：存储需要屏蔽的关键词
    
    def add_blocked_keyword(self, keyword):
        """添加需要屏蔽的关键词"""
        if keyword.strip():
            self.blocked_keywords.add(keyword.strip())
    
    def remove_blocked_keyword(self, keyword):
        """移除屏蔽的关键词"""
        if keyword in self.blocked_keywords:
            self.blocked_keywords.remove(keyword)
    
    def clear_blocked_keywords(self):
        """清空所有屏蔽关键词"""
        self.blocked_keywords.clear()
    
    def build_configs(self, samples, space_required):
        self.level_configs = []
        self.space_required = space_required
        SEPARATORS = '-、.．()（）:：;；·'  # 添加中文点号到分隔符列表
        # 修改分隔符列表，移除顿号
        CONTENT_SEPARATORS = ';；。'  # 基础分隔符
        if self.colon_truncate:
            CONTENT_SEPARATORS += '：:'  # 如果启用冒号截断，添加冒号
        
        for depth, sample in enumerate(samples, 1):
            pattern = ''
            
            # 处理"第X章/部分"这种特殊格式
            m_special = re.match(r'^第(c+|n+)(.*)', sample)
            if m_special:
                type_seq, suffix = m_special.groups()
                if type_seq[0] == 'c':
                    # 修改正则表达式，不再强制要求分隔符
                    pattern = r'^第\s*[一二三四五六七八九十百千万]+\s*' + (re.escape(suffix) if suffix else r'[章节篇部分]?.*')
                elif type_seq[0] == 'n':
                    # 使用非贪婪匹配来处理空格，不再强制要求分隔符
                    pattern = r'^第\s*?\d{1,3}\s*' + (re.escape(suffix) if suffix else r'[章节篇部分]?.*')
            
            # 优先支持括号包裹的c/n/e/cc/nn/ee
            elif re.match(r'^(（+|\(+)(c+|n+|e+|r+)(）+|\)+)', sample):
                m_bracket = re.match(r'^(（+|\(+)(c+|n+|e+|r+)(）+|\)+)', sample)
                left_bracket, type_seq, right_bracket = m_bracket.groups()
                if left_bracket[0] in '（':
                    l_b = '（'
                    r_b = '）'
                else:
                    l_b = '\\('
                    r_b = '\\)'
                
                if type_seq[0] == 'c':
                    # 修改中文数字匹配模式，不再强制要求分隔符
                    pattern = f'^{l_b}[一二三四五六七八九十百千万]+{r_b}.*'
                elif type_seq[0] == 'n':
                    pattern = f'^{l_b}\\d{{1,3}}{r_b}.*'
                elif type_seq[0] == 'e':
                    pattern = f'^{l_b}[a-zA-Z]{{1,3}}{r_b}.*'
                elif type_seq[0] == 'r':
                    # 支持罗马数字(I, II, III, IV等)
                    pattern = f'^{l_b}[IVXLCDM]{{1,4}}{r_b}.*'
            
            # 处理中文点号(·)开头的特殊格式
            elif sample.startswith('·'):
                pattern = r'^·\s*.*'
            
            else:
                # 处理其他格式
                parts = []
                i = 0
                while i < len(sample):
                    if sample[i] == 'c':
                        # 修改中文数字匹配模式，增加对"第"字的支持
                        parts.append('(?:第?\\s*[一二三四五六七八九十百千万]+\\s*[章节篇部分]?|[一二三四五六七八九十百千万]+)')
                        i += 1
                    elif sample[i] == 'n':
                        # 扩展数字匹配，包括可能的空格
                        parts.append('(?:\\s*\\d{1,3}\\s*)')
                        i += 1
                    elif sample[i] == 'e':
                        # 扩展字母匹配，包括可能的空格
                        parts.append('\\s*[a-zA-Z]{1,3}\\s*')
                        i += 1
                    elif sample[i] == 'r':
                        # 支持罗马数字
                        parts.append('[IVXLCDM]{1,4}')
                        i += 1
                    elif sample[i] == '·':
                        # 特殊处理中文点号
                        parts.append('·\\s*')
                        i += 1
                    elif sample[i] in SEPARATORS:
                        parts.append(re.escape(sample[i]))
                        i += 1
                    else:
                        parts.append(re.escape(sample[i]))
                        i += 1
                
                if parts:
                    base_pattern = '^' + ''.join(parts)
                    # 添加标题内容限制，不再强制要求分隔符
                    if space_required[depth-1]:
                        pattern = base_pattern + r'\s+.*'
                    else:
                        pattern = base_pattern + r'.*'
                else:
                    pattern = '^' + re.escape(sample) + r'.*'

            print(f"Level {depth} pattern: {pattern}")
            self.level_configs.append({
                'depth': depth,
                'pattern': re.compile(pattern),
                'max_depth': len(samples)
            })
        self.level_configs.sort(key=lambda x: -x['depth'])
    
    def clean_title(self, title):
        """清理标题，移除页码和多余点号"""
        if not title:
            return title
            
        if not self.remove_page_numbers:
            return title
            
        # 保存原始标题以便调试
        original_title = title
        
        # 首先移除末尾的点号序列
        title = re.sub(r'\.{3,}\s*$', '', title)
        
        # 处理标题后面跟着的页码 - 多种模式，但更保守一些
        # 模式1: 标题...数字  (如 "1.1 反汇编理论..........2")
        title = re.sub(r'\.{3,}\s*\d{1,3}\s*$', '', title)
        
        # 模式2: 标题  数字  (如 "2.1 分类工具  11") - 至少2个空格后跟1-3位数字
        title = re.sub(r'\s{2,}\d{1,3}\s*$', '', title)
        
        # 在冒号处截断（如果启用了该功能）
        if self.colon_truncate:
            # 查找第一个冒号的位置（中文或英文冒号）
            colon_pos = -1
            for i, char in enumerate(title):
                if char in '：:':
                    colon_pos = i
                    break
            if colon_pos >= 0:
                title = title[:colon_pos].strip()
        
        # 检查是否包含屏蔽关键词
        if self.blocked_keywords:
            # 如果整个标题都是屏蔽关键词，返回空字符串
            if title.strip() in self.blocked_keywords:
                return ""
            # 移除标题中的屏蔽关键词
            for keyword in self.blocked_keywords:
                title = title.replace(keyword, "")
            # 清理可能留下的多余空格
            title = " ".join(title.split())
        
        # 如果标题发生变化，打印调试信息
        if title != original_title:
            print(f"清理标题: '{original_title}' -> '{title}'")
        
        return title.strip()
    
    def parse_text(self, text):
        lines = [line.strip() for line in text.split("\n") if line.strip()]
        max_depth = len(self.level_configs)
        outline = []
        current_entry = [""] * max_depth
        last_matched_level = -1  # 记录上一次匹配的层级
        
        # 调试信息
        print(f"\n=== 正在解析 {len(lines)} 行文本 ===")
        print(f"配置的标题格式数量: {len(self.level_configs)}")
        for cfg in self.level_configs:
            print(f"层级 {cfg['depth']} 模式: {cfg['pattern'].pattern}")
        
        # 打印原始行内容
        print("\n=== 原始行内容 ===")
        for i, line in enumerate(lines):
            print(f"行 {i+1}: '{line}'")
        print("=== 原始行内容结束 ===\n")
        
        i = 0
        match_count = 0
        while i < len(lines):
            line = lines[i]
            
            # 检查是否包含屏蔽关键词，如果包含则跳过此行
            if self.blocked_keywords and any(keyword in line for keyword in self.blocked_keywords):
                i += 1
                continue
                
            matched = False
            matched_config = None
            
            # 预处理跨行标题：检测并合并可能的标题行
            # 特别检测三级目录格式 n-n-n
            if re.match(r'^\s*\d+\s*-\s*\d+\s*-\s*\d+\s*$', line) and i+1 < len(lines):
                # 这看起来像是三级目录的编号，但缺少内容部分，尝试合并下一行
                next_line = lines[i+1]
                # 检查下一行是否也是编号行
                next_is_number = False
                for cfg in self.level_configs:
                    if cfg.get('pattern').match(next_line):
                        next_is_number = True
                        break
                
                # 如果下一行不是编号，则合并
                if not next_is_number:
                    print(f"检测到跨行三级标题: '{line}' + '{next_line}'")
                    line = f"{line} {next_line}"
                    i += 1  # 跳过已合并的下一行
            
            # 也检测可能的二级目录格式 n-n
            elif re.match(r'^\s*\d+\s*-\s*\d+\s*$', line) and i+1 < len(lines):
                next_line = lines[i+1]
                next_is_number = False
                for cfg in self.level_configs:
                    if cfg.get('pattern').match(next_line):
                        next_is_number = True
                        break
                
                if not next_is_number:
                    print(f"检测到跨行二级标题: '{line}' + '{next_line}'")
                    line = f"{line} {next_line}"
                    i += 1
            
            # 以及一级目录格式 n-
            elif re.match(r'^\s*\d+\s*-\s*$', line) and i+1 < len(lines):
                next_line = lines[i+1]
                next_is_number = False
                for cfg in self.level_configs:
                    if cfg.get('pattern').match(next_line):
                        next_is_number = True
                        break
                
                if not next_is_number:
                    print(f"检测到跨行一级标题: '{line}' + '{next_line}'")
                    line = f"{line} {next_line}"
                    i += 1
            
            # 尝试匹配当前行
            for config in self.level_configs:
                match = config.get('pattern').match(line)
                if match:
                    match_count += 1
                    depth_idx = config.get('depth') - 1
                    matched = True
                    matched_config = config
                    
                    # 打印匹配信息
                    print(f"匹配成功: 层级 {depth_idx+1}, 行: '{line}'")
                    
                    # 清理标题，移除页码
                    line = self.clean_title(line)
                    
                    # 检查是否需要在分隔符处截断
                    separators = '；;。'  # 基础分隔符
                    if self.colon_truncate:
                        separators += '：:'  # 如果启用冒号截断，添加冒号
                    for sep in separators:
                        sep_idx = line.find(sep)
                        if sep_idx > 0:
                            line = line[:sep_idx].strip()
                            break
                    break
            
            if matched:
                # 检查是否只有编号没有标题内容
                # 通过检查匹配内容是否基本等于整行来判断
                match_content = line.strip()
                is_number_only = True
                for c in match_content:
                    if c.isalpha() and c not in 'IVX':  # 排除罗马数字
                        is_number_only = False
                        break
                
                # 如果只有编号，且还有下一行，尝试合并
                if is_number_only and i + 1 < len(lines):
                    next_line = lines[i + 1]
                    # 确保下一行不是另一个编号
                    next_line_is_number = False
                    for config in self.level_configs:
                        if config.get('pattern').match(next_line):
                            next_line_is_number = True
                            break
                    
                    if not next_line_is_number:
                        # 合并当前行和下一行
                        line = f"{line} {next_line}"
                        line = self.clean_title(line)  # 再次清理合并后的标题
                        i += 1  # 跳过下一行
                
                # 如果是更高层级或同级的新标题，保存当前行并创建新行
                if depth_idx <= last_matched_level:
                    if any(current_entry):
                        outline.append(current_entry[:])
                    current_entry = [""] * max_depth
                    # 保留更高层级的标题
                    for j in range(depth_idx):
                        current_entry[j] = outline[-1][j] if outline else ""
                
                current_entry[depth_idx] = line
                last_matched_level = depth_idx
            
            i += 1

        # 确保最后一行也被添加
        if any(current_entry):
            outline.append(current_entry)

        # 打印匹配结果
        print(f"总共找到 {match_count} 个匹配的标题行")
        print(f"总共生成 {len(outline)} 行大纲数据")
        
        # 移除明显是页码的单独条目，但条件放宽
        filtered_outline = []
        for entry in outline:
            # 只有当所有非空元素都只包含数字时才过滤
            if not all(re.match(r'^\s*\d+\s*$', e) for e in entry if e.strip()):
                filtered_outline.append(entry)

        # 打印原始数据
        print("\n=== 原始提取数据 ===")
        for row in filtered_outline:
            print(row)
        print("=== 原始数据结束 ===\n")

        return filtered_outline

    def _deduplicate(self, outline):
        last_values = {}
        cleaned = []
        for entry in outline:
            new_entry = []
            for idx, value in enumerate(entry):
                if value == last_values.get(idx):
                    new_entry.append("")
                else:
                    new_entry.append(value)
                    last_values[idx] = value
            cleaned.append(new_entry)
        return cleaned

class KeywordDialog(QDialog):
    def __init__(self, parent=None, keywords=None):
        super().__init__(parent)
        self.keywords = keywords or set()
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("关键词屏蔽配置")
        layout = QVBoxLayout()
        
        # 说明标签
        tip_label = QLabel("请输入需要屏蔽的关键词，每行一个：")
        layout.addWidget(tip_label)
        
        # 关键词输入框
        self.keyword_input = QTextEdit()
        self.keyword_input.setPlaceholderText("在此输入关键词，每行一个")
        # 设置现有的关键词
        self.keyword_input.setText("\n".join(sorted(self.keywords)))
        layout.addWidget(self.keyword_input)
        
        # 按钮布局
        btn_layout = QHBoxLayout()
        
        # 确定按钮
        self.ok_button = QPushButton("确定")
        self.ok_button.clicked.connect(self.accept)
        btn_layout.addWidget(self.ok_button)
        
        # 取消按钮
        self.cancel_button = QPushButton("取消")
        self.cancel_button.clicked.connect(self.reject)
        btn_layout.addWidget(self.cancel_button)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
        
        # 设置对话框大小
        self.resize(400, 300)
    
    def get_keywords(self):
        """获取输入的关键词列表"""
        text = self.keyword_input.toPlainText()
        return {line.strip() for line in text.split('\n') if line.strip()}

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.samples = []
        self.space_required = []  # 存储每个层级是否需要空格匹配
        self.extractor = OutlineExtractor()
        
        # 设置应用图标
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pdf.ico')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        
        self.init_ui()
        self.center_window()
        # 添加一个计时器来延迟执行刷新操作
        QTimer.singleShot(100, self.refresh_tables_layout)
    
    def center_window(self):
        # 获取屏幕尺寸
        screen = QApplication.primaryScreen().geometry()
        # 设置固定宽度为屏幕宽度的5/6
        new_width = int(screen.width() * 0.6)
        self.setFixedWidth(new_width)
        # 调整窗口大小
        self.adjustSize()
        # 获取窗口大小
        size = self.size()
        # 计算中心位置
        x = (screen.width() - size.width()) // 2
        y = (screen.height() - size.height()) // 2
        # 移动到屏幕中央
        self.move(x, y)
    
    def init_ui(self):
        self.setWindowTitle("PDF 目录提取器")
        
        # 主布局
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # 配置输入区
        config_group = QWidget()
        config_layout = QVBoxLayout()
        config_layout.setSpacing(5)
        config_layout.setContentsMargins(0, 0, 0, 0)  # 设置边距为0
        
        # 样本输入
        input_label = QLabel("目录层级配置：")
        input_label.setStyleSheet("font-weight: bold;")
        self.sample_input = QTextEdit()
        self.sample_input.setPlaceholderText("请输入从一级目录开始的匹配样本，每行一个\nn：数字 | c：中文 | e：英文 |  r：罗马数字 | 、.-(;)（；）：间隔符  | *：任意内容，例如：\nc、\t\t\t\t\t\t可以代表汉字加顿号\nn.n\t\t\t\t\t\t代表数字点数字\n按 Ctrl + 回车 或点击按钮添加")
        self.sample_input.setMaximumHeight(80)
        self.sample_input.installEventFilter(self)
        
        # 样本操作按钮区域
        sample_btn_layout = QHBoxLayout()
        sample_btn_layout.setSpacing(10)
        self.btn_add = QPushButton("添加目录样本")
        self.btn_add.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_add.setStyleSheet("padding: 5px 15px;")
        self.btn_add.clicked.connect(self.add_samples)
        sample_btn_layout.addWidget(self.btn_add)
        
        # 样本展示区域
        sample_list_container = QWidget()
        sample_list_layout = QHBoxLayout()
        sample_list_layout.setSpacing(5)
        
        # 样本列表
        self.sample_list = QTableWidget()
        self.sample_list.setColumnCount(3)
        self.sample_list.setHorizontalHeaderLabels(["目录层级", "编号结尾强制匹配空格(修改状态后刷新结果)", "正则表达式"])
        self.sample_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        
        # 初始平分宽度
        header = self.sample_list.horizontalHeader()
        total_width = self.sample_list.viewport().width()
        # 20:60:20 的比例分配
        first_column_width = int(total_width * 0.2)  # 20%
        second_column_width = int(total_width * 0.4)  # 40%
        third_column_width = int(total_width * 0.4)  # 40%
        header.setSectionResizeMode(0, QHeaderView.Interactive)
        header.setSectionResizeMode(1, QHeaderView.Interactive)
        header.setSectionResizeMode(2, QHeaderView.Interactive)
        header.resizeSection(0, first_column_width)
        header.resizeSection(1, second_column_width)
        header.resizeSection(2, third_column_width)
        self.sample_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.sample_list.customContextMenuRequested.connect(self.show_context_menu)
        self.sample_list.setMaximumHeight(150)
        self.sample_list.setStyleSheet("QTableWidget::item { padding: 5px; }")
        
        # 排序按钮容器
        sort_btn_container = QWidget()
        sort_btn_container.setFixedWidth(25)
        sort_btn_layout = QVBoxLayout(sort_btn_container)
        sort_btn_layout.setContentsMargins(0, 0, 0, 0)
        sort_btn_layout.setSpacing(0)
        
        # 上下按钮
        self.btn_move_up = QPushButton("↑")
        self.btn_move_down = QPushButton("↓")
        self.btn_move_up.clicked.connect(lambda: self.move_samples(-1))
        self.btn_move_down.clicked.connect(lambda: self.move_samples(1))
        
        # 设置按钮样式和大小策略
        for btn in (self.btn_move_up, self.btn_move_down):
            btn.setStyleSheet("""
                QPushButton {
                    padding: 2px;
                    font-size: 14px;
                    font-weight: bold;
                }
            """)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # 添加按钮到布局
        sort_btn_layout.addWidget(self.btn_move_up)
        sort_btn_layout.addWidget(self.btn_move_down)
        
        # 将列表和按钮添加到容器
        sample_list_layout.addWidget(self.sample_list)
        sample_list_layout.addWidget(sort_btn_container)
        sample_list_container.setLayout(sample_list_layout)

        # 添加所有组件到主布局
        config_layout.addWidget(input_label)
        config_layout.addWidget(self.sample_input)
        config_layout.addLayout(sample_btn_layout)
        config_layout.addWidget(sample_list_container)
        config_group.setLayout(config_layout)

        layout.addWidget(config_group)
        
        # 添加提示标签和文件选择按钮
        file_tip_label = QLabel("选择 PDF 提取目录：")
        file_tip_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(file_tip_label)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        self.btn_file = QPushButton("选择 PDF 文件，自动提取标题")
        self.btn_file.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_file.setStyleSheet("padding: 5px 15px;")
        self.btn_file.clicked.connect(self.select_file)
        btn_layout.addWidget(self.btn_file)
        layout.addLayout(btn_layout)

        # 添加进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("处理进度：%p%")
        self.progress_bar.hide()  # 初始隐藏
        layout.addWidget(self.progress_bar)

        # 分割复选框为左右两组
        left_options = QHBoxLayout()
        right_options = QHBoxLayout()
        
        # 左侧选项（原有的选项）
        self.merge_checkbox = QCheckBox("合并单元格")
        self.merge_checkbox.setChecked(False)
        self.merge_checkbox.stateChanged.connect(self.on_merge_checkbox_changed)
        left_options.addWidget(self.merge_checkbox)
        
        self.vertical_center_checkbox = QCheckBox("合并后竖直居中")
        self.vertical_center_checkbox.setChecked(False)
        self.vertical_center_checkbox.setEnabled(False)
        self.vertical_center_checkbox.stateChanged.connect(self.on_vertical_center_changed)
        left_options.addWidget(self.vertical_center_checkbox)

        self.auto_width_checkbox = QCheckBox("自适应列宽")
        self.auto_width_checkbox.setChecked(False)
        self.auto_width_checkbox.stateChanged.connect(self.on_auto_width_changed)
        left_options.addWidget(self.auto_width_checkbox)
        
        self.remove_page_checkbox = QCheckBox("移除页码")
        self.remove_page_checkbox.setChecked(True)
        self.remove_page_checkbox.stateChanged.connect(self.on_remove_page_changed)
        left_options.addWidget(self.remove_page_checkbox)

        # 新增：冒号截断选项
        self.colon_truncate_checkbox = QCheckBox("冒号处截断")
        self.colon_truncate_checkbox.setChecked(True)
        self.colon_truncate_checkbox.stateChanged.connect(self.on_colon_truncate_changed)
        left_options.addWidget(self.colon_truncate_checkbox)
        
        left_options.addStretch()

        # 右侧选项（OCR相关和关键词配置）
        right_options = QHBoxLayout()
        
        # 添加关键词配置按钮，设置固定宽度
        self.keyword_config_btn = QPushButton("关键词屏蔽配置")
        self.keyword_config_btn.setToolTip("配置需要屏蔽的关键词")
        self.keyword_config_btn.clicked.connect(self.show_keyword_config)
        self.keyword_config_btn.setMinimumWidth(110)  # 设置最小宽度
        right_options.addWidget(self.keyword_config_btn)
        
        if HAS_TESSERACT:            
            self.ocr_settings_btn = QPushButton("OCR 设置")
            self.ocr_settings_btn.setToolTip("配置Tesseract路径和选项")
            self.ocr_settings_btn.clicked.connect(self.show_ocr_settings)
            right_options.addWidget(self.ocr_settings_btn)

            self.force_ocr_checkbox = QCheckBox("强制 OCR 识别")
            self.force_ocr_checkbox.setChecked(False)
            self.force_ocr_checkbox.setToolTip("对所有页面进行OCR识别，适用于扫描版PDF")
            self.force_ocr_checkbox.stateChanged.connect(self.on_force_ocr_changed)
            right_options.addWidget(self.force_ocr_checkbox)

        # 创建底部选项容器
        bottom_options = QHBoxLayout()
        bottom_options.addLayout(left_options)
        bottom_options.addStretch()
        bottom_options.addLayout(right_options)

        # 结果显示
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(1)
        self.result_table.setHorizontalHeaderLabels(["待选择PDF并提取目录、设置显示格式后，可双击单元格编辑提取结果，拖动或点击单个或按住 Ctrl 多选单元格以删除对应行。"])
        self.result_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        # 设置初始列宽
        viewport_width = self.calculate_table_viewport_width()
        if viewport_width > 0:
            self.result_table.setColumnWidth(0, viewport_width - 20)  # 减去滚动条宽度
        self.result_table.setStyleSheet("QTableWidget::item { padding: 5px; }")
        self.result_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.result_table.customContextMenuRequested.connect(self.show_result_context_menu)
        self.result_table.itemDoubleClicked.connect(self.edit_result_item)

        # 保存按钮
        self.btn_save = QPushButton("保存到 Excel")
        self.btn_save.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_save.setStyleSheet("padding: 5px 15px;")
        self.btn_save.clicked.connect(self.save_to_excel)
        self.btn_save.setEnabled(False)

        # 添加所有组件到主布局
        layout.addLayout(bottom_options)
        layout.addWidget(self.result_table)
        layout.addWidget(self.btn_save)
        
        main_widget.setLayout(layout)
    
    def calculate_table_viewport_width(self):
        """计算表格视口的可用宽度"""
        if not hasattr(self, 'result_table'):
            return 0
        # 获取主窗口宽度
        main_width = self.width()
        # 考虑窗口边距和布局间距
        available_width = main_width - 40  # 减去左右边距
        return available_width

    def show_ocr_settings(self):
        """显示OCR设置对话框"""
        # 获取当前路径
        current_path = pytesseract.pytesseract.tesseract_cmd
        
        # 询问用户输入
        path, ok = QInputDialog.getText(
            self, "OCR设置", 
            "请输入Tesseract-OCR安装路径:",
            QLineEdit.Normal,
            current_path
        )
        
        if ok and path:
            # 验证路径
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd = path
                QMessageBox.information(self, "成功", "Tesseract路径已更新")
            else:
                QMessageBox.warning(self, "路径错误", 
                    f"路径 '{path}' 不存在\n请检查Tesseract是否安装")
    
    def eventFilter(self, obj, event):
        if obj is self.sample_input and event.type() == event.KeyPress:
            if event.key() == Qt.Key_Return and event.modifiers() == Qt.ControlModifier:
                self.add_samples()
                return True
        return super().eventFilter(obj, event)
    
    def add_samples(self):
        text = self.sample_input.toPlainText().strip()
        if not text:
            QMessageBox.warning(self, "输入错误", "请输入有效的目录样本")
            return
        
        # 按行分割输入文本
        new_samples = [line.strip() for line in text.split('\n') if line.strip()]
        if not new_samples:
            return
        
        # 添加新样本
        self.samples.extend(new_samples)
        # 为新样本添加空格匹配标志（默认不需要空格）
        self.space_required.extend([False] * len(new_samples))
        self.extractor.build_configs(self.samples, self.space_required)
        
        # 更新样本列表
        self.update_sample_list()
    
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择PDF文件", "", "PDF文件 (*.pdf)")
        if file_path:
            self.current_file = file_path
            # 清除之前提取的文本
            if hasattr(self, 'extracted_text'):
                delattr(self, 'extracted_text')
            # 自动开始提取目录
            self.extract_outline()
    
    def extract_outline(self):
        """提取目录"""
        if not hasattr(self, 'current_file'):
            QMessageBox.warning(self, "错误", "请先选择PDF文件")
            return
        
        try:
            # 显示进度条
            self.progress_bar.show()
            self.progress_bar.setValue(0)
            
            # 禁用所有控件，防止用户操作
            self.setEnabled(False)
            QApplication.processEvents()
            
            try:
                # 如果还没有提取过文本，则提取文本
                if not hasattr(self, 'extracted_text'):
                    # 使用PyMuPDF提取文本，更好地支持多栏结构
                    text = self.extract_text_with_pymupdf(self.current_file)
                    self.extracted_text = text
                
                # 解析目录
                outline = self.extractor.parse_text(self.extracted_text)
                
                # 显示结果
                self.show_results(outline)
                
            finally:
                # 完成后隐藏进度条并重新启用控件
                self.progress_bar.hide()
                self.setEnabled(True)
            
        except Exception as e:
            self.progress_bar.hide()
            self.setEnabled(True)
            QMessageBox.critical(self, "错误", f"提取失败：{str(e)}")
    
    def extract_text_with_pymupdf(self, pdf_path):
        """使用PyMuPDF提取PDF文本，支持多栏结构"""
        try:
            doc = fitz.open(pdf_path)
            text = ""
            total_pages = len(doc)
            
            # 检查是否开启强制OCR模式
            force_ocr = hasattr(self, 'force_ocr_checkbox') and self.force_ocr_checkbox.isChecked()
            
            # 更新进度条标题
            if force_ocr and HAS_TESSERACT:
                self.progress_bar.setFormat("OCR识别PDF中: %p%")
            else:
                self.progress_bar.setFormat("提取文本中: %p%")
            
            for i, page in enumerate(doc):
                # 获取页面尺寸信息
                width, height = page.rect.width, page.rect.height
                
                # 根据模式选择提取方法
                if not force_ocr:
                    # 正常模式：直接提取文本
                    page_text = page.get_text("text")
                    
                    # 检测是否需要OCR (如果页面没有文本或文本极少)
                    if len(page_text.strip()) < 20 and HAS_TESSERACT:
                        try:
                            # 尝试OCR处理
                            # 对大页面降低DPI
                            if width * height > 1000000:  # 超过100万平方点
                                matrix = fitz.Matrix(150/72, 150/72)  # 使用较低DPI
                            else:
                                matrix = fitz.Matrix(300/72, 300/72)  # 默认300 DPI
                                
                            pix = page.get_pixmap(matrix=matrix)
                            img = Image.open(io.BytesIO(pix.tobytes()))
                            
                            # 处理超大图像
                            img_width, img_height = img.size
                            if img_width * img_height > 20000000:  # 2千万像素
                                scale = min(1.0, 4000 / max(img_width, img_height))
                                new_width = int(img_width * scale)
                                new_height = int(img_height * scale)
                                img = img.resize((new_width, new_height), Image.LANCZOS)
                            
                            # 使用中文+英文识别，提高准确率
                            ocr_text = pytesseract.image_to_string(
                                img, 
                                lang='chi_sim+eng',
                                config='--psm 1 --oem 3'  # 自动页面分割，使用LSTM引擎
                            )
                            
                            if ocr_text and len(ocr_text.strip()) > len(page_text.strip()):
                                page_text = ocr_text
                                print(f"第{i+1}页使用OCR结果，识别到{len(ocr_text.strip())}个字符")
                        except Exception as e:
                            print(f"第{i+1}页OCR处理失败: {e}")
                else:
                    # 强制OCR模式：对每一页使用OCR
                    if HAS_TESSERACT:
                        try:
                            # 更新进度条
                            self.progress_bar.setFormat(f"OCR识别第{i+1}/{total_pages}页: %p%")
                            QApplication.processEvents()
                            
                            # 对特别大的页面使用较低DPI
                            if width * height > 1000000:  # 超过100万平方点
                                matrix = fitz.Matrix(150/72, 150/72)  # 使用低DPI
                                print(f"页面{i+1}较大({width:.0f}x{height:.0f})，使用低DPI(150)")
                            else:
                                matrix = fitz.Matrix(300/72, 300/72)  # 默认300 DPI
                            
                            # 获取图像
                            pix = page.get_pixmap(matrix=matrix)
                            img = Image.open(io.BytesIO(pix.tobytes()))
                            
                            # 处理超大图像
                            img_width, img_height = img.size
                            if img_width * img_height > 20000000:  # 2千万像素
                                scale = min(1.0, 3000 / max(img_width, img_height))
                                new_width = int(img_width * scale)
                                new_height = int(img_height * scale)
                                print(f"图像过大({img_width}x{img_height})，缩小至{new_width}x{new_height}")
                                img = img.resize((new_width, new_height), Image.LANCZOS)
                            
                            # 可选：图像预处理
                            try:
                                # 对图像进行增强，提高OCR识别率
                                if img.mode != 'RGB':
                                    img = img.convert('L')  # 转为灰度
                                    
                                    # 使用PIL进行图像增强
                                    from PIL import ImageFilter, ImageEnhance
                                    
                                    # 锐化
                                    img = img.filter(ImageFilter.SHARPEN)
                                    
                                    # 增强对比度
                                    enhancer = ImageEnhance.Contrast(img)
                                    img = enhancer.enhance(2.0)
                                    
                                    # 保存处理后的图像用于调试
                                    debug_dir = os.path.join(os.path.dirname(pdf_path), "debug_ocr")
                                    os.makedirs(debug_dir, exist_ok=True)
                                    debug_file = os.path.join(debug_dir, f"page_{i+1}.png")
                                    img.save(debug_file)
                                    print(f"已保存处理后图像: {debug_file}")
                            except Exception as e:
                                print(f"图像增强失败: {e}")
                            
                            # 尝试多种OCR配置
                            best_text = ""
                            best_len = 0
                            ocr_configs = [
                                '--psm 1 --oem 3',  # 自动分页
                                '--psm 6 --oem 3',  # 单文本块
                            ]
                            
                            for config in ocr_configs:
                                try:
                                    temp_text = pytesseract.image_to_string(
                                        img, 
                                        lang='chi_sim+eng',
                                        config=config
                                    )
                                    if len(temp_text.strip()) > best_len:
                                        best_text = temp_text
                                        best_len = len(temp_text.strip())
                                except Exception as e:
                                    print(f"OCR配置 {config} 失败: {e}")
                                    
                            # 使用最佳结果
                            page_text = best_text if best_len > 0 else "OCR识别失败"
                            print(f"第{i+1}页OCR识别完成，识别到{best_len}个字符")
                        except Exception as e:
                            page_text = f"第{i+1}页OCR处理失败: {e}"
                            print(page_text)
                    else:
                        # 没有安装pytesseract，使用普通提取
                        page_text = page.get_text("text")
                        if not page_text.strip():
                            page_text = f"[第{i+1}页没有识别到文本，请安装pytesseract启用OCR]"
                
                # 添加页码信息
                text += f"=== 第{i+1}页 ===\n{page_text}\n"
                
                # 更新进度条
                progress = int((i + 1) / total_pages * 100)
                self.progress_bar.setValue(progress)
                QApplication.processEvents()
            
            # 如果提取的文本太少且不是强制OCR模式，尝试备用方法
            if len(text.strip()) < 100 and not force_ocr:
                # 备用方法：使用pdfplumber
                return self.extract_text_with_pdfplumber(pdf_path)
            
            # 重置进度条格式
            self.progress_bar.setFormat("处理进度：%p%")
            
            return text
        except Exception as e:
            print(f"PyMuPDF提取失败: {e}")
            # 回退到pdfplumber
            return self.extract_text_with_pdfplumber(pdf_path)
    
    def extract_text_with_pdfplumber(self, pdf_path):
        """使用pdfplumber提取PDF文本(备用方法)"""
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            text = ""
            for i, page in enumerate(pdf.pages):
                try:
                    # 尝试按表格提取，这可能有助于保持多栏结构
                    tables = page.extract_tables()
                    if tables:
                        # 处理表格
                        for table in tables:
                            for row in table:
                                text += " | ".join([cell if cell else "" for cell in row]) + "\n"
                    
                    # 再提取普通文本
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                except Exception as e:
                    print(f"提取页面 {i+1} 时出错: {e}")
                    # 尝试基本提取
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                    except:
                        pass
                
                # 更新进度条
                progress = int((i + 1) / total_pages * 100)
                self.progress_bar.setValue(progress)
                QApplication.processEvents()
            
            text = f"=== 第{i+1}页 ===\n{text}\n"
            return text

    def show_context_menu(self, pos):
        menu = QMenu(self)
        delete_action = menu.addAction("删除选中项")
        clear_action = menu.addAction("清空所有")
        
        action = menu.exec_(self.sample_list.mapToGlobal(pos))
        
        if action == delete_action:
            self.delete_selected_samples()
        elif action == clear_action:
            self.clear_all_samples()
    
    def delete_selected_samples(self):
        selected_rows = set(item.row() for item in self.sample_list.selectedItems())
        if not selected_rows:
            return
            
        # 从后向前删除，避免索引变化
        for row in sorted(selected_rows, reverse=True):
            del self.samples[row]
        
        # 更新界面
        self.update_sample_list()
    
    def clear_all_samples(self):
        self.samples.clear()
        self.space_required.clear()
        self.update_sample_list()
    
    def move_samples(self, direction):
        current_row = self.sample_list.currentRow()
        if current_row == -1:
            return
            
        new_row = current_row + direction
        if 0 <= new_row < len(self.samples):
            # 交换样本位置
            self.samples[current_row], self.samples[new_row] = \
                self.samples[new_row], self.samples[current_row]
            # 更新界面
            self.update_sample_list()
            # 选中移动后的行
            self.sample_list.selectRow(new_row)
    
    def update_sample_list(self):
        self.sample_list.setRowCount(len(self.samples))
        # 确保space_required列表长度与samples相同
        while len(self.space_required) < len(self.samples):
            self.space_required.append(False)
        
        # 重新构建配置以获取正则表达式
        self.extractor.build_configs(self.samples, self.space_required)
        
        # 移除旧的信号连接
        try:
            self.sample_list.itemChanged.disconnect()
            self.sample_list.itemClicked.disconnect()
        except:
            pass
        
        for row, sample in enumerate(self.samples):
            # 目录层级
            level_item = QTableWidgetItem(f"{convert_to_chinese_num(row+1)}级目录")
            level_item.setTextAlignment(Qt.AlignCenter)  # 居中对齐
            self.sample_list.setItem(row, 0, level_item)
            
            # 空格匹配复选框
            checkbox_text = "强制" if self.space_required[row] else "不强制"
            checkbox_item = QTableWidgetItem(checkbox_text)
            checkbox_item.setTextAlignment(Qt.AlignCenter)  # 居中对齐
            checkbox_item.setFlags(Qt.ItemIsEnabled)  # 允许点击但不允许编辑文本
            if self.space_required[row]:
                checkbox_item.setBackground(Qt.lightGray)  # 选中时背景色变化
            self.sample_list.setItem(row, 1, checkbox_item)
            
            # 正则表达式
            pattern = self.extractor.level_configs[-(row+1)]['pattern'].pattern
            pattern_item = QTableWidgetItem(pattern)
            pattern_item.setToolTip(pattern)  # 添加工具提示，方便查看完整表达式
            pattern_item.setFlags(pattern_item.flags() | Qt.ItemIsEditable)  # 允许编辑
            self.sample_list.setItem(row, 2, pattern_item)
        
        # 连接新的信号
        self.sample_list.itemClicked.connect(self.on_item_clicked)
        self.sample_list.itemChanged.connect(self.on_item_changed)
        
        self.btn_save.setEnabled(len(self.samples) > 0)
    
    def on_item_clicked(self, item):
        if item.column() == 1:  # 空格匹配列
            try:
                # 如果正在处理中，直接返回
                if not self.isEnabled():
                    return
                    
                # 显示进度条
                self.show_progress(True, "更新目录匹配规则")
                
                # 禁用所有控件
                self.setEnabled(False)
                self.sample_list.setEnabled(False)  # 特别禁用样本列表
                QApplication.processEvents()
                
                try:
                    # 记录当前的合并状态
                    was_merged = self.merge_checkbox.isChecked()
                    if was_merged:
                        # 临时取消合并状态
                        self.merge_checkbox.setChecked(False)
                        QApplication.processEvents()
                    
                    row = item.row()
                    self.space_required[row] = not self.space_required[row]  # 切换状态
                    # 更新显示
                    checkbox_text = "强制" if self.space_required[row] else "不强制"
                    item.setText(checkbox_text)
                    if self.space_required[row]:
                        item.setBackground(Qt.lightGray)
                    else:
                        item.setBackground(Qt.white)
                    
                    self.update_progress(20)
                    
                    # 重新构建配置
                    self.extractor.build_configs(self.samples, self.space_required)
                    # 更新正则表达式显示
                    pattern = self.extractor.level_configs[-(row+1)]['pattern'].pattern
                    pattern_item = QTableWidgetItem(pattern)
                    pattern_item.setToolTip(pattern)
                    pattern_item.setFlags(pattern_item.flags() | Qt.ItemIsEditable)
                    self.sample_list.setItem(row, 2, pattern_item)
                    
                    self.update_progress(40)
                    
                    # 如果有提取的文本，直接使用现有数据重新解析
                    if hasattr(self, 'extracted_text'):
                        # 使用已有的文本重新解析
                        outline = self.extractor.parse_text(self.extracted_text)
                        self.update_progress(70)
                        # 更新结果显示
                        self.show_results(outline)
                        
                        self.update_progress(90)
                        
                        # 如果之前是合并状态，恢复合并
                        if was_merged:
                            QApplication.processEvents()
                            self.merge_checkbox.setChecked(True)
                    
                    self.update_progress(100)
                finally:
                    # 重新启用所有控件
                    self.setEnabled(True)
                    self.sample_list.setEnabled(True)
                    QApplication.processEvents()
                    # 隐藏进度条
                    self.show_progress(False)
            except Exception as e:
                # 发生错误时确保界面可用
                self.setEnabled(True)
                self.sample_list.setEnabled(True)
                self.show_progress(False)
                QApplication.processEvents()
                QMessageBox.warning(self, "错误", f"处理空格匹配状态时发生错误：{str(e)}")
                # 出错时尝试恢复界面状态
                if hasattr(self, 'original_outline'):
                    self.show_results(self.original_outline)
    
    def on_item_changed(self, item):
        if item.column() == 2:  # 正则表达式列
            try:
                # 尝试编译正则表达式
                re.compile(item.text())
                # 如果编译成功，更新配置
                row = item.row()
                self.extractor.level_configs[-(row+1)]['pattern'] = re.compile(item.text())
                # 如果有PDF文件已经打开，重新提取目录
                if hasattr(self, 'current_file'):
                    self.extract_outline()
            except re.error as e:
                # 如果正则表达式无效，显示错误消息
                QMessageBox.warning(self, "正则表达式错误", f"输入的正则表达式无效：{str(e)}")
                # 恢复原来的正则表达式
                pattern = self.extractor.level_configs[-(item.row()+1)]['pattern'].pattern
                item.setText(pattern)
    
    def show_result_context_menu(self, pos):
        menu = QMenu(self)
        selected_rows = set(item.row() for item in self.result_table.selectedItems())
        
        if len(selected_rows) > 0:
            if len(selected_rows) == 1:
                delete_action = menu.addAction("删除此行")
            else:
                delete_action = menu.addAction(f"删除选中的 {len(selected_rows)} 行")
            
            action = menu.exec_(self.result_table.mapToGlobal(pos))
            
            if action == delete_action:
                # 从后向前删除行，避免索引变化
                for row in sorted(selected_rows, reverse=True):
                    self.result_table.removeRow(row)
    
    def edit_result_item(self, item):
        item.setFlags(item.flags() | Qt.ItemIsEditable)
    
    def show_results(self, outline):
        """显示提取结果"""
        try:
            if not outline:
                return

            # 显示进度条
            self.show_progress(True, "更新显示结果")
            
            # 保存原始数据用于后续操作
            self.original_outline = outline
            # 保存去重后的数据
            self.deduped_outline = self.extractor._deduplicate(outline)
            
            self.update_progress(10)

            # 暂时阻止表格更新以提高性能
            self.result_table.setUpdatesEnabled(False)
            self.sample_list.setUpdatesEnabled(False)
            
            try:
                self.result_table.clear()
                self.result_table.setColumnCount(len(self.samples))
                self.result_table.setHorizontalHeaderLabels([f"{convert_to_chinese_num(i+1)}级目录" for i in range(len(self.samples))])
                
                self.update_progress(20)
                
                # 设置列宽模式
                header = self.result_table.horizontalHeader()
                for i in range(self.result_table.columnCount()):
                    header.setSectionResizeMode(i, QHeaderView.Interactive)
                
                # 设置初始列宽
                if not self.auto_width_checkbox.isChecked():
                    self.adjust_equal_column_widths()
                
                self.update_progress(30)

                if self.merge_checkbox.isChecked():
                    # 使用原始数据进行合并显示
                    self.result_table.setRowCount(len(self.original_outline))
                    total_rows = len(self.original_outline)
                    # 填充数据
                    for row, entry in enumerate(self.original_outline):
                        for col, value in enumerate(entry):
                            item = QTableWidgetItem(value)
                            if self.vertical_center_checkbox.isChecked():
                                item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                            else:
                                item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
                            self.result_table.setItem(row, col, item)
                        # 更新填充数据的进度
                        self.update_progress(30 + int((row / total_rows) * 30))

                    # 处理合并
                    total_cols = self.result_table.columnCount()
                    for col in range(total_cols):
                        row = 0
                        while row < self.result_table.rowCount():
                            # 更新合并进度
                            self.update_progress(60 + int((col / total_cols) * 30))
                            
                            # 获取当前单元格的值
                            current_item = self.result_table.item(row, col)
                            if not current_item:
                                row += 1
                                continue
                            
                            current_value = current_item.text().strip()
                            if not current_value:
                                row += 1
                                continue
                            
                            # 找到可以合并的行范围
                            merge_start = row
                            merge_count = 1
                            next_row = row + 1
                            
                            # 检查后续行是否可以合并
                            while next_row < self.result_table.rowCount():
                                # 检查所有前面的列是否相同
                                can_merge = True
                                for prev_col in range(col):
                                    prev_current = self.result_table.item(merge_start, prev_col)
                                    prev_next = self.result_table.item(next_row, prev_col)
                                    if (prev_current and prev_current.text().strip()) != (prev_next and prev_next.text().strip()):
                                        can_merge = False
                                        break
                                
                                if not can_merge:
                                    break
                                    
                                # 检查当前列的值是否相同
                                next_item = self.result_table.item(next_row, col)
                                next_value = next_item.text().strip() if next_item else ''
                                
                                if next_value == current_value:
                                    merge_count += 1
                                    next_row += 1
                                else:
                                    break
                            
                            # 如果找到可以合并的行
                            if merge_count > 1:
                                try:
                                    # 设置合并
                                    self.result_table.setSpan(merge_start, col, merge_count, 1)
                                    # 设置合并后的单元格样式
                                    merged_item = QTableWidgetItem(current_value)
                                    if self.vertical_center_checkbox.isChecked():
                                        merged_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                                    else:
                                        merged_item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
                                    self.result_table.setItem(merge_start, col, merged_item)
                                except Exception as e:
                                    print(f"设置合并单元格时出错: {str(e)}")
                                
                            row = merge_start + merge_count
                else:
                    # 直接显示去重后的数据
                    self.result_table.setRowCount(len(self.deduped_outline))
                    total_rows = len(self.deduped_outline)
                    for row, entry in enumerate(self.deduped_outline):
                        for col, value in enumerate(entry):
                            item = QTableWidgetItem(value)
                            item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
                            self.result_table.setItem(row, col, item)
                        # 更新进度
                        self.update_progress(60 + int((row / total_rows) * 30))

                # 自动调整行高
                self.adjust_row_heights()
                
                # 恢复表格更新
                self.result_table.setUpdatesEnabled(True)
                self.sample_list.setUpdatesEnabled(True)
                QApplication.processEvents()
                
                # 在表格完全加载后，如果启用了自适应列宽，重新调整列宽
                if self.auto_width_checkbox.isChecked():
                    QTimer.singleShot(100, self.adjust_auto_column_widths)  # 延迟100ms执行
                
                self.update_progress(100)
                
            finally:
                # 确保表格可以更新
                self.result_table.setUpdatesEnabled(True)
                self.sample_list.setUpdatesEnabled(True)
                QApplication.processEvents()
                # 隐藏进度条
                self.show_progress(False)
                
        except Exception as e:
            self.show_progress(False)
            QMessageBox.warning(self, "错误", f"显示结果时出错：{str(e)}")
            # 确保表格可以更新
            self.result_table.setUpdatesEnabled(True)
            self.sample_list.setUpdatesEnabled(True)

    def adjust_equal_column_widths(self):
        """设置等宽的列宽"""
        viewport_width = self.calculate_table_viewport_width()
        if viewport_width > 0:
            available_width = viewport_width - 20  # 减去滚动条宽度
            if self.result_table.columnCount() > 1:
                # 计算网格线占用的总宽度（每个网格线占2像素）
                grid_width = (self.result_table.columnCount() - 1) * 2
                # 计算每列实际可用宽度（使用math.floor确保向下取整）
                import math
                column_width = math.floor((available_width - grid_width) / self.result_table.columnCount())
                # 设置所有列的宽度
                for i in range(self.result_table.columnCount()):
                    self.result_table.setColumnWidth(i, column_width)
            else:
                # 单列情况：使用全部可用宽度
                self.result_table.setColumnWidth(0, available_width)

    def adjust_auto_column_widths(self):
        """根据内容自动调整列宽"""
        for col in range(self.result_table.columnCount()):
            max_width = 0
            # 检查表头宽度
            header_text = self.result_table.horizontalHeaderItem(col).text()
            header_width = self.result_table.fontMetrics().boundingRect(header_text).width() + 20
            max_width = max(max_width, header_width)
            
            # 检查每行该列的内容宽度
            for row in range(self.result_table.rowCount()):
                item = self.result_table.item(row, col)
                if item:
                    content = item.text()
                    # 计算文本宽度，考虑中文字符
                    content_width = 0
                    for char in content:
                        if ord(char) > 127:  # 中文字符
                            content_width += int(self.result_table.fontMetrics().boundingRect(char).width() * 1.1)
                        else:
                            content_width += self.result_table.fontMetrics().boundingRect(char).width()
                    content_width = int(content_width + 20)  # 添加一些padding
                    max_width = max(max_width, content_width)
            
            # 限制最小和最大宽度，确保是整数
            max_width = min(max(int(max_width), 100), 400)
            self.result_table.setColumnWidth(col, max_width)

    def adjust_row_heights(self):
        """自动调整所有行的高度"""
        for row in range(self.result_table.rowCount()):
            # 获取当前行中所有单元格的内容高度
            row_height = 0
            for col in range(self.result_table.columnCount()):
                item = self.result_table.item(row, col)
                if item:
                    # 获取单元格内容的理想高度
                    content = item.text()
                    # 计算文本换行后的高度
                    font_metrics = self.result_table.fontMetrics()
                    col_width = self.result_table.columnWidth(col)
                    text_rect = font_metrics.boundingRect(
                        0, 0, col_width - 10, 1000,  # 减去一些边距
                        Qt.TextWordWrap | Qt.AlignLeft | Qt.AlignTop,
                        content
                    )
                    row_height = max(row_height, text_rect.height() + 10)  # 加上一些边距
            
            # 设置最小行高
            row_height = max(row_height, 30)  # 最小30像素
            self.result_table.setRowHeight(row, row_height)
    
    def on_merge_checkbox_changed(self, state):
        """处理合并单元格复选框状态改变"""
        try:
            # 显示进度条
            self.show_progress(True, "更新表格显示")
            
            # 禁用所有控件，防止用户操作
            self.setEnabled(False)
            QApplication.processEvents()
            
            try:
                # 启用/禁用竖直居中选项
                self.vertical_center_checkbox.setEnabled(state == Qt.Checked)
                
                # 如果没有数据，直接返回
                if not hasattr(self, 'original_outline') or not self.original_outline:
                    return
                
                # 清除表格内容和合并状态
                self.result_table.clearContents()
                
                # 取消所有合并
                for row in range(self.result_table.rowCount()):
                    for col in range(self.result_table.columnCount()):
                        try:
                            if self.result_table.rowSpan(row, col) > 1 or self.result_table.columnSpan(row, col) > 1:
                                # 在取消合并前保存单元格内容
                                current_item = self.result_table.item(row, col)
                                current_text = current_item.text() if current_item else ""
                                current_alignment = current_item.textAlignment() if current_item else (Qt.AlignLeft | Qt.AlignTop)
                                
                                # 取消合并
                                self.result_table.setSpan(row, col, 1, 1)
                                
                                # 恢复单元格内容
                                if current_text:
                                    new_item = QTableWidgetItem(current_text)
                                    new_item.setTextAlignment(current_alignment)
                                    self.result_table.setItem(row, col, new_item)
                        except Exception as e:
                            print(f"取消合并单元格时出错 ({row}, {col}): {str(e)}")
                            continue
                
                # 重新显示数据
                if state == Qt.Checked:
                    # 使用原始数据进行合并显示
                    self.result_table.setRowCount(len(self.original_outline))
                    for row, entry in enumerate(self.original_outline):
                        for col, value in enumerate(entry):
                            item = QTableWidgetItem(value)
                            if self.vertical_center_checkbox.isChecked():
                                item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                            else:
                                item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
                            self.result_table.setItem(row, col, item)
                    
                    # 处理合并
                    for col in range(self.result_table.columnCount()):
                        row = 0
                        while row < self.result_table.rowCount():
                            # 获取当前单元格的值
                            current_item = self.result_table.item(row, col)
                            if not current_item:
                                row += 1
                                continue
                            
                            current_value = current_item.text().strip()
                            if not current_value:
                                row += 1
                                continue
                            
                            # 找到可以合并的行范围
                            merge_start = row
                            merge_count = 1
                            next_row = row + 1
                            
                            # 检查后续行是否可以合并
                            while next_row < self.result_table.rowCount():
                                # 检查所有前面的列是否相同
                                can_merge = True
                                for prev_col in range(col):
                                    prev_current = self.result_table.item(merge_start, prev_col)
                                    prev_next = self.result_table.item(next_row, prev_col)
                                    if (prev_current and prev_current.text().strip()) != (prev_next and prev_next.text().strip()):
                                        can_merge = False
                                        break
                                
                                if not can_merge:
                                    break
                                    
                                # 检查当前列的值是否相同
                                next_item = self.result_table.item(next_row, col)
                                next_value = next_item.text().strip() if next_item else ''
                                
                                if next_value == current_value:
                                    merge_count += 1
                                    next_row += 1
                                else:
                                    break
                            
                            # 如果找到可以合并的行
                            if merge_count > 1:
                                try:
                                    # 设置合并
                                    self.result_table.setSpan(merge_start, col, merge_count, 1)
                                    # 设置合并后的单元格样式
                                    merged_item = QTableWidgetItem(current_value)
                                    if self.vertical_center_checkbox.isChecked():
                                        merged_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                                    else:
                                        merged_item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
                                    self.result_table.setItem(merge_start, col, merged_item)
                                except Exception as e:
                                    print(f"设置合并单元格时出错: {str(e)}")
                            
                            row = merge_start + merge_count
                else:
                    # 显示去重后的数据
                    self.result_table.setRowCount(len(self.deduped_outline))
                    for row, entry in enumerate(self.deduped_outline):
                        for col, value in enumerate(entry):
                            item = QTableWidgetItem(value)
                            item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
                            self.result_table.setItem(row, col, item)
                
                # 调整列宽和行高
                if self.auto_width_checkbox.isChecked():
                    QTimer.singleShot(100, self.adjust_auto_column_widths)
                else:
                    self.adjust_equal_column_widths()
                self.adjust_row_heights()
                
            finally:
                # 重新启用所有控件
                self.setEnabled(True)
                QApplication.processEvents()
                # 隐藏进度条
                self.show_progress(False)
            
        except Exception as e:
            # 确保界面可用
            self.setEnabled(True)
            self.show_progress(False)
            QApplication.processEvents()
            QMessageBox.warning(self, "错误", f"处理合并单元格时出错：{str(e)}")
            # 出错时尝试恢复界面状态
            if hasattr(self, 'original_outline'):
                self.show_results(self.original_outline)

    def on_vertical_center_changed(self, state):
        """处理竖直居中复选框状态改变"""
        try:
            # 如果没有数据，直接返回
            if not hasattr(self, 'original_outline') or not self.original_outline:
                return
                
            # 仅更新单元格的对齐方式，不重新处理合并
            for row in range(self.result_table.rowCount()):
                for col in range(self.result_table.columnCount()):
                    item = self.result_table.item(row, col)
                    if item:
                        if state == Qt.Checked:
                            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                        else:
                            item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"更新单元格对齐方式时出错：{str(e)}")

    def on_auto_width_changed(self, state):
        """处理自适应列宽复选框状态改变"""
        if hasattr(self, 'original_outline'):
            if state == Qt.Checked:
                self.adjust_auto_column_widths()
            else:
                self.adjust_equal_column_widths()
            # 调整完列宽后自动调整行高
            self.adjust_row_heights()

    def on_remove_page_changed(self, state):
        """处理移除页码复选框状态改变"""
        self.extractor.remove_page_numbers = (state == Qt.Checked)
        # 如果已经有提取结果，重新提取
        if hasattr(self, 'extracted_text'):
            self.extract_outline()

    def on_force_ocr_changed(self, state):
        """处理强制OCR复选框状态改变"""
        if state == Qt.Checked:
            # 检查Tesseract是否可用
            if not HAS_TESSERACT:
                QMessageBox.warning(self, "OCR组件缺失", 
                    "未检测到pytesseract库或Tesseract-OCR引擎。\n"
                    "请确保正确安装了pytesseract和Tesseract-OCR，并包含中文语言包。")
                self.force_ocr_checkbox.setChecked(False)
                return
                
            # 检查Tesseract是否正确配置
            try:
                if not os.path.exists(pytesseract.pytesseract.tesseract_cmd):
                    path_text = pytesseract.pytesseract.tesseract_cmd
                    
                    # 尝试常见安装路径
                    common_paths = [
                        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"
                    ]
                    
                    found = False
                    for path in common_paths:
                        if os.path.exists(path):
                            pytesseract.pytesseract.tesseract_cmd = path
                            found = True
                            break
                    
                    if not found:
                        QMessageBox.warning(self, "Tesseract路径错误", 
                            f"当前Tesseract路径无效: {path_text}\n"
                            "请通过「OCR设置」按钮设置正确的Tesseract-OCR路径。")
                        self.show_ocr_settings()
                        
                        # 如果还是不正确，禁用强制OCR
                        if not os.path.exists(pytesseract.pytesseract.tesseract_cmd):
                            self.force_ocr_checkbox.setChecked(False)
                            return
            except Exception as e:
                QMessageBox.warning(self, "Tesseract配置错误", f"错误: {str(e)}")
                self.force_ocr_checkbox.setChecked(False)
                return
            
            # 检查中文语言包
            try:
                langs = pytesseract.get_languages()
                if 'chi_sim' not in langs:
                    QMessageBox.warning(self, "缺少中文语言包", 
                        "Tesseract-OCR未安装中文语言包(chi_sim)。\n"
                        "请到Tesseract官网下载中文语言包并放置到tessdata目录。")
            except Exception as e:
                print(f"检查语言包时出错: {e}")
            
            # 提示用户OCR可能需要时间
            QMessageBox.information(self, "OCR处理", 
                "强制OCR模式已启用，扫描版PDF处理可能需要较长时间，请耐心等待。\n"
                "OCR识别进度将在状态栏显示。")
            
            # 禁用页码移除（避免可能的错误）
            if self.remove_page_checkbox.isChecked():
                self.remove_page_checkbox.setChecked(False)
            
            # 如果已经有提取结果，重新提取
            if hasattr(self, 'extracted_text'):
                # 清除之前的结果，强制重新OCR
                delattr(self, 'extracted_text')
                self.extract_outline()
        else:
            # 关闭强制OCR，恢复正常模式
            if hasattr(self, 'extracted_text'):
                # 清除之前的结果，使用普通方式重新提取
                delattr(self, 'extracted_text')
                self.extract_outline()

    def on_colon_truncate_changed(self, state):
        """处理冒号截断复选框状态改变"""
        self.extractor.colon_truncate = (state == Qt.Checked)
        # 如果已经有提取结果，重新提取
        if hasattr(self, 'extracted_text'):
            self.extract_outline()

    def refresh_tables_layout(self):
        """刷新两个表格的布局"""
        # 刷新sample_list的列宽
        header = self.sample_list.horizontalHeader()
        total_width = self.sample_list.viewport().width()
        # 考虑右侧按钮的宽度(大约25px)和一些边距(5px)
        available_width = total_width - 15
        if available_width > 0:
            # 20:30:50 的比例分配
            first_column_width = int(available_width * 0.2)  # 20%
            second_column_width = int(available_width * 0.4)  # 40%
            third_column_width = available_width - first_column_width - second_column_width  # 60%
            header.setSectionResizeMode(0, QHeaderView.Interactive)
            header.setSectionResizeMode(1, QHeaderView.Interactive)
            header.setSectionResizeMode(2, QHeaderView.Interactive)
            header.resizeSection(0, first_column_width)
            header.resizeSection(1, second_column_width)
            header.resizeSection(2, third_column_width)
        
        # 刷新result_table的列宽
        header = self.result_table.horizontalHeader()
        total_width = self.result_table.viewport().width()
        column_count = self.result_table.columnCount()
        if column_count > 0:  # 确保有列才进行平分
            column_width = total_width // column_count
            for i in range(column_count):
                header.setSectionResizeMode(i, QHeaderView.Interactive)
                header.resizeSection(i, column_width)
    
    def resizeEvent(self, event):
        """窗口大小改变时触发"""
        super().resizeEvent(event)
        # 窗口大小改变时，延迟刷新表格布局
        QTimer.singleShot(100, self.refresh_tables_layout)
        
        # 根据是否自适应列宽来设置宽度
        if self.auto_width_checkbox.isChecked():
            self.adjust_auto_column_widths()
        else:
            self.adjust_equal_column_widths()

    def save_to_excel(self):
        save_path, _ = QFileDialog.getSaveFileName(
            self, "保存结果", "", "Excel文件 (*.xlsx)")
        if save_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active

                # 写入表头
                headers = [f"{convert_to_chinese_num(i+1)}级目录" for i in range(len(self.samples))]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # 从表格中获取当前显示的数据
                for row in range(self.result_table.rowCount()):
                    for col in range(self.result_table.columnCount()):
                        # Excel的行列号从1开始，且要考虑表头行
                        excel_row = row + 2
                        excel_col = col + 1
                        
                        # 获取单元格内容
                        item = self.result_table.item(row, col)
                        if item is not None:
                            value = item.text()
                            cell = ws.cell(row=excel_row, column=excel_col, value=value)
                            
                            # 设置对齐方式
                            if self.merge_checkbox.isChecked():
                                if self.vertical_center_checkbox.isChecked():
                                    cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='left')
                                else:
                                    cell.alignment = openpyxl.styles.Alignment(vertical='top', horizontal='left')
                            else:
                                cell.alignment = openpyxl.styles.Alignment(vertical='top', horizontal='left')

                # 如果需要合并单元格
                if self.merge_checkbox.isChecked():
                    # 遍历每一列
                    for col in range(self.result_table.columnCount()):
                        row = 0
                        while row < self.result_table.rowCount():
                            # 获取合并信息
                            rowspan = self.result_table.rowSpan(row, col)
                            colspan = self.result_table.columnSpan(row, col)
                            
                            # 如果是合并单元格
                            if rowspan > 1 or colspan > 1:
                                # Excel的行列号从1开始，且要考虑表头行
                                start_row = row + 2
                                start_col = col + 1
                                end_row = start_row + rowspan - 1
                                end_col = start_col + colspan - 1
                                
                                # 合并单元格
                                ws.merge_cells(
                                    start_row=start_row,
                                    start_column=start_col,
                                    end_row=end_row,
                                    end_column=end_col
                                )
                                
                                # 设置合并后的单元格样式
                                merged_cell = ws.cell(row=start_row, column=start_col)
                                if self.vertical_center_checkbox.isChecked():
                                    merged_cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='left')
                                else:
                                    merged_cell.alignment = openpyxl.styles.Alignment(vertical='top', horizontal='left')
                            
                            # 移动到下一个未合并的单元格
                            row += max(rowspan, 1)

                # 自动调整列宽
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:  # 只处理有值的单元格
                                cell_length = 0
                                for char in str(cell.value):
                                    if ord(char) > 127:  # 中文字符
                                        cell_length += 2
                                    else:
                                        cell_length += 1
                                max_length = max(max_length, cell_length)
                        except:
                            pass
                    adjusted_width = max_length + 4
                    ws.column_dimensions[column].width = adjusted_width

                # 保存文件
                wb.save(save_path)

                # 显示成功对话框
                msg_box = QMessageBox()
                msg_box.setWindowTitle("成功")
                msg_box.setText(f"文件已保存到：{save_path}")
                
                open_file_btn = msg_box.addButton("打开文件", QMessageBox.ActionRole)
                open_folder_btn = msg_box.addButton("打开文件夹", QMessageBox.ActionRole)
                close_btn = msg_box.addButton("关闭", QMessageBox.RejectRole)
                
                msg_box.exec_()
                
                clicked_button = msg_box.clickedButton()
                if clicked_button == open_file_btn:
                    os.startfile(save_path)
                elif clicked_button == open_folder_btn:
                    os.startfile(os.path.dirname(save_path))
                
            except Exception as e:
                QMessageBox.critical(self, "错误", f"保存文件时出错：{str(e)}")

    def show_keyword_config(self):
        """显示关键词配置对话框"""
        dialog = KeywordDialog(self, self.extractor.blocked_keywords)
        if dialog.exec_() == QDialog.Accepted:
            # 获取新的关键词列表
            new_keywords = dialog.get_keywords()
            # 更新提取器的关键词列表
            self.extractor.blocked_keywords = new_keywords
            # 如果已经有提取结果，重新提取
            if hasattr(self, 'extracted_text'):
                self.extract_outline()

    def show_progress(self, show=True, text="处理中"):
        """显示或隐藏进度条
        Args:
            show: 是否显示进度条
            text: 进度条显示的文本
        """
        if show:
            self.progress_bar.setFormat(f"{text}: %p%")
            self.progress_bar.setValue(0)
            self.progress_bar.show()
        else:
            self.progress_bar.hide()
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("处理进度：%p%")
    
    def update_progress(self, value, maximum=100):
        """更新进度条
        Args:
            value: 当前进度值
            maximum: 最大进度值
        """
        progress = int((value / maximum) * 100)
        self.progress_bar.setValue(progress)
        QApplication.processEvents()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # 设置应用程序级别的图标
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pdf.ico')
    if os.path.exists(icon_path):
        app.setWindowIcon(QIcon(icon_path))
    
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
